"""
plot_PDTs.py

Reads PDTs.xlsx and generates two horizontal bar charts:
    e) # Unique PDTs per cell line and combined
    f) # Kinases with PDTs per cell line and combined

Run from the project root:
    python plot_PDTs.py
"""

import openpyxl
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

# ── Load data ─────────────────────────────────────────────────────────────────

wb = openpyxl.load_workbook("PDTs.xlsx", read_only=True)

pdts_per_line   = {}
kinases_per_line = {}

for cell_line in ["MCF7", "HL60", "NTERA2"]:
    ws = wb[cell_line]
    unique_pdts  = set()
    kinase_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        kinase, n, substrates = row[0], row[1], row[2]
        if n and int(n) > 0:
            kinase_count += 1
        if substrates:
            for site in substrates.split(";"):
                site = site.strip()
                if site:
                    unique_pdts.add(site)
    pdts_per_line[cell_line]    = unique_pdts
    kinases_per_line[cell_line] = kinase_count

all_pdts    = pdts_per_line["MCF7"] | pdts_per_line["HL60"] | pdts_per_line["NTERA2"]
all_kinases = len(kinases_per_line["MCF7"] and kinases_per_line or {})

# combined kinases: union of kinase names across sheets
wb2 = openpyxl.load_workbook("PDTs.xlsx", read_only=True)
combined_kinases = set()
for cell_line in ["MCF7", "HL60", "NTERA2"]:
    ws = wb2[cell_line]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            combined_kinases.add(row[0])

# ── Colours ───────────────────────────────────────────────────────────────────

COLORS = {
    "Combined": "#4CAF50",   # green
    "NTERA2":   "#E53935",   # red
    "MCF7":     "#5C6BC0",   # purple
    "HL60":     "#FFA000",   # amber
}

# ── Helper: horizontal bar chart ──────────────────────────────────────────────

def bar_chart(ax, labels, values, colors, xlabel, title):
    bars = ax.barh(labels, values, color=colors, height=0.55, edgecolor="none")
    ax.set_xlabel(xlabel, fontsize=10)
    ax.set_title(title, fontsize=11, fontweight="bold", loc="left")
    ax.invert_yaxis()                       # Combined on top, HL60 at bottom
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.set_xlim(0, max(values) * 1.18)

    # Value labels inside a white box at the right end of each bar
    for bar, val in zip(bars, values):
        ax.text(
            val + max(values) * 0.01,
            bar.get_y() + bar.get_height() / 2,
            f"{val:,}",
            va="center", ha="left", fontsize=9,
            bbox=dict(boxstyle="round,pad=0.2", fc="white", ec="gray", lw=0.6),
        )

# ── Build figure ──────────────────────────────────────────────────────────────

fig, (ax_e, ax_f) = plt.subplots(1, 2, figsize=(10, 3.5))
fig.subplots_adjust(wspace=0.45)

# Panel e — Unique PDTs
labels_e = ["Combined", "NTERA2", "MCF7", "HL60"]
values_e = [
    len(all_pdts),
    len(pdts_per_line["NTERA2"]),
    len(pdts_per_line["MCF7"]),
    len(pdts_per_line["HL60"]),
]
colors_e = [COLORS[l] for l in labels_e]
bar_chart(ax_e, labels_e, values_e, colors_e,
          xlabel="# Unique PDTs", title="e    # Unique PDTs")
ax_e.set_ylabel("Cell lines", fontsize=10)

# Panel f — Kinases with PDTs
labels_f = ["Combined", "NTERA2", "MCF7", "HL60"]
values_f = [
    len(combined_kinases),
    kinases_per_line["NTERA2"],
    kinases_per_line["MCF7"],
    kinases_per_line["HL60"],
]
colors_f = [COLORS[l] for l in labels_f]
bar_chart(ax_f, labels_f, values_f, colors_f,
          xlabel="# Kinases with PDTs", title="f    # Kinases with PDTs")

plt.savefig("PDTs_summary.png", dpi=150, bbox_inches="tight")
plt.savefig("PDTs_summary.svg", bbox_inches="tight")
plt.show()

print("\nResults:")
print(f"  Unique PDTs — MCF7: {len(pdts_per_line['MCF7'])}, "
      f"HL60: {len(pdts_per_line['HL60'])}, "
      f"NTERA2: {len(pdts_per_line['NTERA2'])}, "
      f"Combined: {len(all_pdts)}")
print(f"  Kinases    — MCF7: {kinases_per_line['MCF7']}, "
      f"HL60: {kinases_per_line['HL60']}, "
      f"NTERA2: {kinases_per_line['NTERA2']}, "
      f"Combined: {len(combined_kinases)}")
print("\nSaved: PDTs_summary.png  /  PDTs_summary.svg")

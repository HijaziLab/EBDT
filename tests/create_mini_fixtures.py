"""
create_mini_fixtures.py : Generates HL60_mini.xlsm, NTERA2_mini.xlsm and MCF7_mini.xlsm tests/fixtures/.

Takes the first 20 rows of data (+ header) of the sheets 'fold' and 'pvalue'
of each real .xlsm

run from the project root:
    python tests/create_mini_fixtures.py
"""

import os
import openpyxl

# ==== ROUTE ====

CODE_DIR = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
FIXTURES_DIR = os.path.join(os.path.dirname(__file__), "fixtures")

N_ROWS = 20  #without counting the header

# ==== MAIN FUNCTION =====

def create_mini(cell_line):
    src_path = os.path.join(CODE_DIR, f"{cell_line}.xlsm")
    dst_path = os.path.join(FIXTURES_DIR, f"{cell_line}_mini.xlsm")

    print(f"Reading {src_path}...")
    src_wb = openpyxl.load_workbook(src_path, read_only=True, keep_vba=False)

    dst_wb = openpyxl.Workbook()
    dst_wb.remove(dst_wb.active)  # delete the black sheet by default

    for sheet_name in ["fold", "pvalue"]:
        src_ws = src_wb[sheet_name]
        dst_ws = dst_wb.create_sheet(sheet_name)

        rows = list(src_ws.iter_rows(values_only=True))
        # header + first N_ROWS of data
        for row in rows[:N_ROWS + 1]:
            dst_ws.append(list(row))

        print(f"  {sheet_name}: {len(rows[:N_ROWS + 1])} rows, {len(rows[0])} cols")

    src_wb.close()

    dst_wb.save(dst_path)
    print(f"Saved: {dst_path}\n")


# ==== RUN =====

for cell_line in ["MCF7", "HL60", "NTERA2"]:
    create_mini(cell_line)

print("Mini fixtures created correctly.")

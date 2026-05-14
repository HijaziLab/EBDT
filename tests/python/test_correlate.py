"""
test_correlate.py

unitary tests for correlatePhosphoPeptideWithInhibitorSpecificity().

This function does 2 independent things:
  1. CORRELATION: for each kinase with >=2 compounds, calculates the Pearson 
     correlation between inhibition values and fold changes.
  2. RATIO: calculate what fraction of compounds significantly inhibit
     each phosphosite (fold<-1 and p-value<0.025).

test data used in this module:

KinaseX: inhibits CompA and CompB : has calculated correlation
KinaseY: inhibits only CompA    : no correlation (needs >=2 compounds)

SITE1: both fold<-1 and p<0.025 : ratio=1.0 for KinaseX
SITE2: both fold>-1 : ratio=0 for KinaseX

Correlation KinaseX-SITE1=+1.0 (calculated by hand):
  kinaseValues=[0.01, 0.05], foldValues=[-2.5, -1.5]
  Both increase together : r=+1.0

Run:
    pytest tests/python/test_correlate.py -v
"""

import pytest
import openpyxl
from conftest import make_workbook
from ebdtFunctions import correlatePhosphoPeptideWithInhibitorSpecificity


# ==== DATA TEST ====

KINASES = {
    "KinaseX": (["CompA", "CompB"], [0.01, 0.05]),  # 2 compounds = valid
    "KinaseY": (["CompA"],          [0.02]),         # 1 compound  = no correlation
}

F_VALUES = {
    "SITE1;": {"CompA": -2.5, "CompB": -1.5},   # both<-1 - ratio=1.0
    "SITE2;": {"CompA":  0.8, "CompB":  0.5},   # both>-1 - ratio=0
}

P_VALUES = {
    "SITE1;": {"CompA": 0.010, "CompB": 0.010}, # significant (<0.025)
    "SITE2;": {"CompA": 0.800, "CompB": 0.900}, # no significant
}

COMPOUNDS_KUSTER = {
    "CompA": "SK6.CompA",
    "CompB": "JAK2.CompB",
}

COMPOUNDS_CELL_LINE = {
    "CompA": "MCF7.CompA.fold",
    "CompB": "MCF7.CompB.fold",
}


@pytest.fixture
def empty_wb():
    #empty workbook (function needs oneto write intermediate sheets)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


@pytest.fixture
def ratios(empty_wb):
    #Runs function and returns calculated ratios
    return correlatePhosphoPeptideWithInhibitorSpecificity(
        empty_wb, KINASES, F_VALUES, P_VALUES, COMPOUNDS_KUSTER, COMPOUNDS_CELL_LINE
    )


# ==== TESTS: RATIO ====

def test_ratio_uno_cuando_todos_compuestos_inhiben(ratios):
    """
    SITE1 has fold<-1 and p<0.025 for both compounds of KinaseX.
    Ratio=2 compounds that inhibits/2 total compounds= 1.0
    """
    assert ratios["KinaseX"]["SITE1;"] == pytest.approx(1.0)


def test_ratio_cero_cuando_fold_no_baja_de_menos_uno(ratios):
    """
    SITE2 has fold>-1 for all compounds.
    The fold chnge treshold is -1, so none of them counts : ratio=0.
    """
    assert ratios["KinaseX"]["SITE2;"] == pytest.approx(0.0)


def test_ratio_cero_cuando_pvalue_no_significativo():
    """
    even if the fold is <-1, if p-value>=0.025 the compound does not count.
    """
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    kinases = {"KinaseX": (["CompA", "CompB"], [0.01, 0.05])}
    f = {"SITE1;": {"CompA": -3.0, "CompB": -2.0}}   # fold ok
    p = {"SITE1;": {"CompA":  0.1, "CompB":  0.5}}   # p-value no significant

    ratios = correlatePhosphoPeptideWithInhibitorSpecificity(
        wb, kinases, f, p, {}, {}
    )
    assert ratios["KinaseX"]["SITE1;"] == pytest.approx(0.0)


def test_ratio_parcial_cuando_solo_un_compuesto_inhibe():

    #if only 1 of 2 compounds meets the requirement (fold<-1 and p<0.025): ratio=1/2=0.5
    
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    kinases = {"KinaseX": (["CompA", "CompB"], [0.01, 0.05])}
    f = {"SITE1;": {"CompA": -2.0, "CompB":  0.5}}   # only CompA<-1
    p = {"SITE1;": {"CompA":  0.01, "CompB": 0.01}}  # both significant

    ratios = correlatePhosphoPeptideWithInhibitorSpecificity(
        wb, kinases, f, p, {}, {}
    )
    assert ratios["KinaseX"]["SITE1;"] == pytest.approx(0.5)


def test_kinasa_con_un_compuesto_tiene_ratio_dict_vacio(ratios):
    """
    KinaseY only has 1 compound. Code requires >=2 to calculate ratio.
    the KinaseY ratio dict must be empty.
    """
    assert ratios["KinaseY"] == {}


# === TESTS: CORRELACION ====

def test_correlacion_positiva_cuando_ambos_aumentan(empty_wb):
    """
    kinaseValues=[0.01, 0.05] and foldValues=[-2.5, -1.5]: both increase.
    Pearson correlation with 2 aligned points=+1.0
    """
    kinases = {"KinaseX": (["CompA", "CompB"], [0.01, 0.05])}
    f = {"SITE1;": {"CompA": -2.5, "CompB": -1.5}}
    p = {"SITE1;": {"CompA": 0.5, "CompB": 0.5}}

    correlatePhosphoPeptideWithInhibitorSpecificity(
        empty_wb, kinases, f, p, {}, {}
    )
    # Verification of the correlation on the written sheet
    ws = empty_wb["corrPPwithKinases"]
    rows = list(ws.iter_rows(values_only=True))
    # Row 1=Kinase headers, Col 1=p-sites, data from [1][1]
    corr_val = rows[1][1]  # row SITE1, column KinaseX
    assert corr_val == pytest.approx(1.0, abs=1e-9)


def test_correlacion_negativa_cuando_uno_decrece(empty_wb):
    """
    kinaseValues=[0.01, 0.05] and foldValues=[0.8, 0.5]: inhibition increases, fold goes down.
    Correlation = -1.0
    """
    kinases = {"KinaseX": (["CompA", "CompB"], [0.01, 0.05])}
    f = {"SITE2;": {"CompA": 0.8, "CompB": 0.5}}
    p = {"SITE2;": {"CompA": 0.5, "CompB": 0.5}}

    correlatePhosphoPeptideWithInhibitorSpecificity(
        empty_wb, kinases, f, p, {}, {}
    )
    ws = empty_wb["corrPPwithKinases"]
    rows = list(ws.iter_rows(values_only=True))
    corr_val = rows[1][1]
    assert corr_val == pytest.approx(-1.0, abs=1e-9)


def test_correlacion_cero_cuando_un_solo_compuesto(empty_wb):
    #Using only one compound cannot calculate correlation : must be 0.
    kinases = {"KinaseY": (["CompA"], [0.02])}
    f = {"SITE1;": {"CompA": -2.5}}
    p = {"SITE1;": {"CompA": 0.01}}

    correlatePhosphoPeptideWithInhibitorSpecificity(
        empty_wb, kinases, f, p, {}, {}
    )
    ws = empty_wb["corrPPwithKinases"]
    # KinaseY cannot appear because it has <2 compounds
    header_row = list(ws.iter_rows(values_only=True))[0]
    assert "KinaseY" not in header_row


# ==== TESTS: HOJAS CREADAS ====

def test_crea_hoja_corrPPwithKinases(ratios, empty_wb):
    #Function must create the sheet 'corrPPwithKinases'
    assert "corrPPwithKinases" in empty_wb.sheetnames


def test_crea_hoja_ratioSigPPoverSignInhSpeci(ratios, empty_wb):
    #function must create sheet 'ratioSigPPoverSignInhSpeci'
    assert "ratioSigPPoverSignInhSpeci" in empty_wb.sheetnames

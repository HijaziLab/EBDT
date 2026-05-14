"""
test_createNodeEdges.py

unit test for createNodeEdges().

This function finds kinase pairs that share substrates and creates 2 sheets in workbook:

  - 'tableEdgeSubs': simmetrical matrix NxN. Cell [i,j] contains shared substrates
    between kinase i and kinase j, separated by ';'.

  - 'nodes.edges': list of pairs with columns:
        edge   = "KinasaA.KinasaB"
        weight = number of shared substrates
        subs   = substrates separated by ';'

Run:
    pytest tests/python/test_createNodeEdges.py -v
"""

import pytest
import openpyxl
from ebdtFunctions import createNodeEdges


# ==== TEST DATA ====

# KinaseA y KinaseB share SITE1 and SITE3 : must appear in nodes.edges
# KinaseC only has SITE4 : nothing is shared : doesn`t appear in edges
KINASE_SUBSTRATES = {
    "KinaseA": ["SITE1", "SITE2", "SITE3"],
    "KinaseB": ["SITE1", "SITE3"],
    "KinaseC": ["SITE4"],
}


@pytest.fixture
def empty_wb():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


@pytest.fixture
def wb_con_edges(empty_wb):
    #Runs createNodeEdges and returns the resulting workbook.
    createNodeEdges(empty_wb, KINASE_SUBSTRATES)
    return empty_wb


# ==== TESTS: nodes.edges ====

def test_par_con_sustratos_compartidos_aparece_en_edges(wb_con_edges):
    #KinaseA and KinaseB shere SITE1 and SITE3 : must appear in nodes.edges.
    ws = wb_con_edges["nodes.edges"]
    edges = [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    assert "KinaseA.KinaseB" in edges


def test_par_sin_sustratos_compartidos_no_aparece_en_edges(wb_con_edges):
    #KinaseA and KinaseC do not share substrates : cannot appear in nodes.edges.
    ws = wb_con_edges["nodes.edges"]
    edges = [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    assert "KinaseA.KinaseC" not in edges
    assert "KinaseB.KinaseC" not in edges


def test_weight_correcto(wb_con_edges):
    #weight of KinaseA-KinaseB must be 2 (SITE1 and SITE3 are shared).
    ws = wb_con_edges["nodes.edges"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == "KinaseA.KinaseB":
            assert row[1] == 2
            return
    pytest.fail("Cannot find the pair KinaseA.KinaseB in nodes.edges")


def test_sustratos_compartidos_en_edges(wb_con_edges):
    #Column 'subs' must contains shared substrates.
    ws = wb_con_edges["nodes.edges"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == "KinaseA.KinaseB":
            subs = set(row[2].split(";"))
            assert "SITE1" in subs
            assert "SITE3" in subs
            return
    pytest.fail("Cannot find the pair KinaseA.KinaseB in nodes.edges")


# ==== TESTS: tableEdgeSubs ====

def test_tabla_tableEdgeSubs_tiene_dimensiones_correctas(wb_con_edges):
    #tableEdgeSubs must be a matrix: (n_quinasas+1) x (n_quinasas+1).
    ws = wb_con_edges["tableEdgeSubs"]
    rows = list(ws.iter_rows(values_only=True))
    n = len(KINASE_SUBSTRATES)
    assert len(rows) == n + 1      # +1 for the headers row
    assert len(rows[0]) == n + 1   # +1 for the headers column


def test_tabla_tableEdgeSubs_es_simetrica(wb_con_edges):
    """
    tableEdgeSubs[i,j] must be equal to tableEdgeSubs[j,i].
    shared substrates table is simmetrycal by definition.
    """
    ws = wb_con_edges["tableEdgeSubs"]
    rows = list(ws.iter_rows(values_only=True))
    n = len(KINASE_SUBSTRATES)

    for i in range(1, n + 1):
        for j in range(1, n + 1):
            val_ij = rows[i][j] or ""
            val_ji = rows[j][i] or ""
            assert val_ij == val_ji, (
                f"Asymetry in [{i},{j}]: '{val_ij}' != '{val_ji}'"
            )


# ==== TESTS: CREATED SHEETS ====

def test_crea_hoja_tableEdgeSubs(wb_con_edges):
    #Function must create the sheet 'tableEdgeSubs'
    assert "tableEdgeSubs" in wb_con_edges.sheetnames


def test_crea_hoja_nodes_edges(wb_con_edges):
    #Function must create the sheet 'nodes.edges'
    assert "nodes.edges" in wb_con_edges.sheetnames


def test_nodes_edges_tiene_header_correcto(wb_con_edges):
    #sheet nodes.edges must have headers: 'edge', 'weight', 'subs'
    ws = wb_con_edges["nodes.edges"]
    header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    assert header[0] == "edge"
    assert header[1] == "weight"
    assert header[2] == "subs"


#==== TEST: BORDERLINE CASE: no shared substrates ====

def test_sin_sustratos_compartidos_nodes_edges_vacio():
    #if no kinase share substrates, nodes.edges only has a header
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    substrates = {
        "KinaseA": ["SITE1"],
        "KinaseB": ["SITE2"],  #no overlap
    }
    createNodeEdges(wb, substrates)
    ws = wb["nodes.edges"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data_rows = [r for r in rows if any(v is not None for v in r)]
    assert data_rows == []

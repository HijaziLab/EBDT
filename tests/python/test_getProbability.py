"""
test_getProbability.py

unit test for getProbabilityofBeingKinaseSubs().

This function calaculates the probability that each psite is a substrate 
of each kinase, normalizing the ratio by the maximun observed ratio:
    prob(psite, kinase)=ratio(psite, kinase)/max_ratio(psite)
Key rules:
  - if kinase is not expressed in cell line: prob=0 for all its sites
  - if maxRatio=0 for a psite: prob=0 (protection against divide by 0)
  - the maximun possible probability is 1.0

test Data:
KinaseX expressed in MCF7, KinaseY not expressed.
  SITE1: ratio KinaseX=1.0 : prob=1.0/1.0= 1.0
  SITE2: ratio KinaseX=0.0 : prob=0.0/0.0 : pprotected: 0.0

Run:
    pytest tests/python/test_getProbability.py -v
"""

import pytest
import openpyxl
from ebdtFunctions import getProbabilityofBeingKinaseSubs


# ==== TEST DATA =====

# calculated ratios by correlatePhosphoPeptideWithInhibitorSpecificity
RATIOS = {
    "KinaseX": {"SITE1;": 1.0, "SITE2;": 0.0},
    "KinaseY": {},  # 1 compound : empty dict 
}

# expressed kinases by cell line (format: string separated by ';')
LIST_OF_KINASES = {
    "MCF7": "KinaseX;PRKCA;CDK7",   #expresed KinaseX , KinaseY not
}


@pytest.fixture
def empty_wb():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


@pytest.fixture
def probs(empty_wb):
    #runs the fuction and returns probability dictionary
    return getProbabilityofBeingKinaseSubs(empty_wb, LIST_OF_KINASES, RATIOS, "MCF7")


# ==== TESTS: EXPRESSED KINASE ====

def test_probabilidad_uno_cuando_ratio_es_maximo(probs):
    """
    SITE1 has ratio=1.0 for KinaseX, that is the maximun.
    prob=1.0/1.0=1.0
    """
    assert probs["KinaseX"]["SITE1;"] == pytest.approx(1.0)


def test_probabilidad_cero_cuando_ratio_es_cero(probs):
    """
    SITE2 has ratio=0.0, and maxRatio for SITE2 is 0.0 too.
    Code protects division : prob=0
    """
    assert probs["KinaseX"]["SITE2;"] == pytest.approx(0.0)


def test_probabilidades_entre_0_y_1(probs):
    #all probabilities should be in range [0.0, 1.0]
    for kinase, site_dict in probs.items():
        for site, prob in site_dict.items():
            assert 0.0 <= prob <= 1.0, (
                f"Probabilities out of range: probs[{kinase}][{site}] = {prob}"
            )


# ==== TESTS: UNEXPRESSED KINASE ====

def test_probabilidad_cero_para_quinasa_no_expresada(probs):
    """
    KinaseY is not in the kinase list of MCF7
    all its probabilities must be 0
    (in this case its dict is empty because its ratio dict is empty too)
    """
    assert probs["KinaseY"] == {}


def test_quinasa_no_expresada_con_sitios_tiene_prob_cero():
    """
    if a kinase has ratios (was processed) but it is not expressed
    in cell line, its probabilities must be all 0.
    """
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    ratios = {
        "KinaseNoExpresada": {"SITE1;": 0.8, "SITE2;": 0.5},
    }
    list_of_kinases = {"MCF7": "OtraKinasa;PRKCA"}  # KinaseNoExpresada missing

    probs = getProbabilityofBeingKinaseSubs(wb, list_of_kinases, ratios, "MCF7")

    assert probs["KinaseNoExpresada"]["SITE1;"] == pytest.approx(0.0)
    assert probs["KinaseNoExpresada"]["SITE2;"] == pytest.approx(0.0)


# ==== TESTS: BORDERLINE CASE - maxRatio=0 ====

def test_maxRatio_cero_no_produce_division_por_cero():
    """
    if all the ratios of a psite are 0, maxRatio=0
    code shound manage this without raissing a ZeroDivisionError
    """
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    ratios = {
        "KinaseX": {"SITE1;": 0.0, "SITE2;": 0.0},  #all 0
    }
    list_of_kinases = {"MCF7": "KinaseX"}

    # should not raise exception
    probs = getProbabilityofBeingKinaseSubs(wb, list_of_kinases, ratios, "MCF7")

    assert probs["KinaseX"]["SITE1;"] == pytest.approx(0.0)
    assert probs["KinaseX"]["SITE2;"] == pytest.approx(0.0)


def test_normalizacion_correcta_con_varios_ratios():
    """
    with multiple kinases and maxRatio>0, normalizaction must be correct
    maxRatio of SITE1=max(0.8, 0.4)=0.8
    prob KinaseA=0.8/0.8=1.0
    prob KinaseB=0.4/0.8=0.5
    """
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    ratios = {
        "KinaseA": {"SITE1;": 0.8},
        "KinaseB": {"SITE1;": 0.4},
    }
    list_of_kinases = {"MCF7": "KinaseA;KinaseB"}

    probs = getProbabilityofBeingKinaseSubs(wb, list_of_kinases, ratios, "MCF7")

    assert probs["KinaseA"]["SITE1;"] == pytest.approx(1.0)
    assert probs["KinaseB"]["SITE1;"] == pytest.approx(0.5)


# ==== TESTS: CREATED SHEETS ====
def test_crea_hoja_ProbOfBeingKinaseSubs(probs, empty_wb):
    #function must create the sheet 'ProbOfBeingKinaseSubs'
    assert "ProbOfBeingKinaseSubs" in empty_wb.sheetnames

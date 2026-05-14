"""
test_integration.py

integration test: runs the full pipeline with mini data
(fixtures) and verify that the result has the correct structure

These tests do not verify exact numbers (this is a job for 
test_golden.py), but they verify:
  - pipeline finishes without errors
  - all expected sheets are generated
  - sheets have the correct number of columns and rows
  - data types are coherent

They need the mini fixtures in tests/fixtures/.

run:
    pytest tests/python/test_integration.py -v
"""

import os
import sys
import pytest
import openpyxl

# conftest.py already adds the CODE_DIR to the sys.path
from ebdtFunctions import GetExpectancyOfBeingDownstreamTarget


# ==== EXPECTED SHEETS IN OUTPUT EXCEL =====

HOJAS_INPUT  = {"pvalue", "fold"}
HOJAS_OUTPUT = {
    "pvalue.select",
    "fold.select",
    "corrPPwithKinases",
    "ratioSigPPoverSignInhSpeci",
    "ProbOfBeingKinaseSubs",
    "PutativeKinaseSubstrates",
    "tableEdgeSubs",
    "nodes.edges",
}
TODAS_LAS_HOJAS = HOJAS_INPUT | HOJAS_OUTPUT


# ==== FIXTURE: EXECUTED PIPELINE ====

@pytest.fixture(scope="module", params=["MCF7", "HL60", "NTERA2"])
def wb_resultado(request, tmp_path_factory):
    """
    Runs the full pipeline with the mini fixture for each cell line and returns
    the resulting workbook. Parametrized: runs once per cell line (MCF7, HL60, NTERA2).

    Uses scope='module' so each pipeline only executes once for all tests of this file.
    """
    import shutil
    from conftest import FIXTURES_DIR
    cell_line = request.param
    run_dir = str(tmp_path_factory.mktemp(f"{cell_line.lower()}_integration_run"))

    shutil.copy(
        os.path.join(FIXTURES_DIR, f"{cell_line}_mini.xlsm"),
        os.path.join(run_dir, f"{cell_line}.xlsm")
    )
    shutil.copytree(
        os.path.join(FIXTURES_DIR, "requiredData"),
        os.path.join(run_dir, "requiredData")
    )

    old_cwd = os.getcwd()
    try:
        os.chdir(run_dir)
        GetExpectancyOfBeingDownstreamTarget(
            inhibitionThreshold  = 0.1,
            ratioThreshold       = 0.2,
            probabilityThreshold = 0.5,
            cellLinesFiles       = [f"{cell_line}.xlsm"],
        )
    finally:
        os.chdir(old_cwd)

    return openpyxl.load_workbook(
        os.path.join(run_dir, f"{cell_line}.xlsm"),
        read_only=True,
        keep_vba=False
    )


# ==== TESTS: SHEETS ====

def test_se_generan_todas_las_hojas_de_output(wb_resultado):
    #pipeline must generate 8 sheets of output
    sheets = set(wb_resultado.sheetnames)
    for hoja in HOJAS_OUTPUT:
        assert hoja in sheets, f"missing sheet: '{hoja}'"


def test_hojas_de_input_se_conservan(wb_resultado):
    #'fold' and 'pvalue' original sheets must be preserved
    for hoja in HOJAS_INPUT:
        assert hoja in wb_resultado.sheetnames


# ==== TESTS: PutativeKinaseSubstrates ====

def test_PutativeKinaseSubstrates_tiene_header_correcto(wb_resultado):
    #the first row of PutativeKinaseSubstrates must be: kinase, n, substrates.
    ws = wb_resultado["PutativeKinaseSubstrates"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    assert header[0] == "kinase"
    assert header[1] == "n"
    assert header[2] == "substrates"


def test_PutativeKinaseSubstrates_tiene_una_fila_por_quinasa(wb_resultado):
    
    #It must be at least one row of kinase, but not more than the total of the CSV (163).
    
    ws = wb_resultado["PutativeKinaseSubstrates"]
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    data_rows = [r for r in data_rows if r[0] is not None]
    assert len(data_rows) >= 1
    assert len(data_rows) <= 163  


def test_PutativeKinaseSubstrates_columna_n_es_numerica(wb_resultado):
    #Column 'n' (number of substrates) must contains integers
    ws = wb_resultado["PutativeKinaseSubstrates"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is not None:
            assert isinstance(row[1], (int, float)), (
                f"Column 'n' is not numerical: {row[1]!r}"
            )


# ==== TESTS: nodes.edges ====

def test_nodes_edges_tiene_header_correcto(wb_resultado):
    #nodes.edges must start with: edge, weight, subs."""
    ws = wb_resultado["nodes.edges"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    assert header[0] == "edge"
    assert header[1] == "weight"
    assert header[2] == "subs"


# ==== TESTS: corrPPwithKinases ====

def test_corrPPwithKinases_tiene_valores_numericos(wb_resultado):
    """
    data cells of corrPPwithKinases must be numerical
    (Pearson correlations in [-1, 1]) or None if they were not calculated.
    """
    ws = wb_resultado["corrPPwithKinases"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows[:5]:  # verify the first 5 rows
        for val in row[1:]:
            if val is not None:
                assert isinstance(val, (int, float)), f"No numerical value: {val!r}"
                assert -1.01 <= val <= 1.01, f"out of range correlation: {val}"


# ==== TESTS: ratioSigPPoverSignInhSpeci ====

def test_ratios_entre_0_y_1(wb_resultado):
    #all ratios must be in [0, 1]
    ws = wb_resultado["ratioSigPPoverSignInhSpeci"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        for val in row[1:]:
            if val is not None:
                assert 0 <= val <= 1, f"out of range ratio [0,1]: {val}"


# ==== TESTS: ProbOfBeingKinaseSubs ====

def test_probabilidades_entre_0_y_1(wb_resultado):
    #all probabilities must be in [0, 1]
    ws = wb_resultado["ProbOfBeingKinaseSubs"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        for val in row[1:]:
            if val is not None:
                assert 0 <= val <= 1, f"out of range probability [0,1]: {val}"


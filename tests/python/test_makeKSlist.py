"""
test_makeKSlist.py

unit test for makeKSlistOfKinaseDownstreamTargets().

This function filters psites and determines which are the putative substrates of each kinase
For a psite to be included, 3 conditions must meet simultaneously
    1. prob >= probabilityThreshold
    2. ratio >= ratioThreshold
    3. fdr < 0.02  (fixed threshold in the code)

Moreover, psites that contain the next residues will be excluded:
    'None', '(M', '(R', '(K'

and the multi-sites psites ('SITE1;SITE2;') will be divided into individual sites

run:
    pytest tests/python/test_makeKSlist.py -v
"""

import pytest
import openpyxl
from ebdtFunctions import makeKSlistOfKinaseDownstreamTargets


# ==== BASE DATA ====

# Probabilities: KinaseX has high prob in SITE1, low in SITE2
PROB_KINASES = {
    "KinaseX": {"SITE1;": 1.0, "SITE2;": 0.0},
    "KinaseY": {},
}

# Ratios
RATIOS = {
    "KinaseX": {"SITE1;": 1.0, "SITE2;": 0.0},
    "KinaseY": {},
}

# FDR: SITE1 pass the filter (< 0.02), SITE2 do not
FDR_VALUES = {
    "SITE1;": 0.010,
    "SITE2;": 0.500,
}

# P-values (required by the function, but nos used directy here)
P_VALUES = {
    "SITE1;": {"CompA": 0.01},
    "SITE2;": {"CompA": 0.80},
}


@pytest.fixture
def empty_wb():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


def _run(wb, probs=None, ratios=None, fdr=None, ratio_t=0.5, prob_t=0.5):
    #runs makeKSlistOfKinaseDownstreamTargets with values by default
    return makeKSlistOfKinaseDownstreamTargets(
        wb,
        ratios  or RATIOS,
        probs   or PROB_KINASES,
        P_VALUES,
        fdr     or FDR_VALUES,
        ratioThreshold       = ratio_t,
        probabilityThreshold = prob_t,
    )


# ==== TESTS: TRIPLE FILTER ====

def test_sitio_incluido_cuando_supera_triple_umbral(empty_wb):
    """
    SITE1: prob=1.0>=0.5, ratio=1.0>=0.5, fdr=0.01<0.02 : must appear
    in the substrate list of KinaseX.
    """
    result = _run(empty_wb)
    assert "SITE1" in result["KinaseX"]


def test_sitio_incluido_en_el_umbral_exacto(empty_wb):
    """
    Boundary test: prob == probabilityThreshold and ratio == ratioThreshold
    must be INCLUDED (>= is non-strict, matching VBA behaviour).
    """
    probs  = {"KinaseX": {"SITE1;": 0.5}}   # exactly at threshold
    ratios = {"KinaseX": {"SITE1;": 0.5}}   # exactly at threshold
    fdr    = {"SITE1;": 0.001}

    result = _run(empty_wb, probs=probs, ratios=ratios, fdr=fdr, prob_t=0.5, ratio_t=0.5)
    assert "SITE1" in result["KinaseX"]


def test_sitio_excluido_cuando_fdr_alto(empty_wb):
    #SITE2: fdr=0.5>=0.02 : excluded even prob and ratio are high
    
    probs = {"KinaseX": {"SITE2;": 1.0}}
    ratios = {"KinaseX": {"SITE2;": 1.0}}
    fdr = {"SITE2;": 0.5}  #high FDR : excluded

    result = _run(empty_wb, probs=probs, ratios=ratios, fdr=fdr)
    assert "SITE2" not in result["KinaseX"]


def test_sitio_excluido_cuando_prob_baja(empty_wb):
    #prob=0.3<0.5 (probabilityThreshold) : excluded
    
    probs = {"KinaseX": {"SITE1;": 0.3}}
    ratios = {"KinaseX": {"SITE1;": 1.0}}
    fdr = {"SITE1;": 0.01}

    result = _run(empty_wb, probs=probs, ratios=ratios, fdr=fdr, prob_t=0.5)
    assert "SITE1" not in result["KinaseX"]


def test_sitio_excluido_cuando_ratio_bajo(empty_wb):
    #ratio=0.1<0.5 (ratioThreshold) : excluded
    
    probs = {"KinaseX": {"SITE1;": 1.0}}
    ratios = {"KinaseX": {"SITE1;": 0.1}}
    fdr = {"SITE1;": 0.01}

    result = _run(empty_wb, probs=probs, ratios=ratios, fdr=fdr, ratio_t=0.5)
    assert "SITE1" not in result["KinaseX"]


# ==== TESTS: EXCLUSION OF RESIDUES ====

@pytest.mark.parametrize("sitio_con_residuo", [
    "PROT(M1);",    # contains (M
    "PROT(R5);",    # contains (R
    "PROT(K10);",   # contains (K
    "None",         
])
def test_excluye_fosfositios_con_residuos_especiales(sitio_con_residuo, empty_wb):
    """
    psites that contain '(M', '(R', '(K' o 'None' must be excluded,
    independently of its prob, ratio and fdr.
    """
    probs  = {"KinaseX": {sitio_con_residuo: 1.0}}
    ratios = {"KinaseX": {sitio_con_residuo: 1.0}}
    fdr    = {sitio_con_residuo: 0.001}

    result = _run(empty_wb, probs=probs, ratios=ratios, fdr=fdr)
    assert result["KinaseX"] == []


# ==== TESTS: MULTI-SITE ====

def test_split_multisite_en_sitios_individuales(empty_wb):
    """
    A psite like 'SITE1;SITE2;' must be divided into ['SITE1', 'SITE2']
    Empty elements (by final ';') must be ignored
    """
    probs  = {"KinaseX": {"SITE1;SITE2;": 1.0}}
    ratios = {"KinaseX": {"SITE1;SITE2;": 1.0}}
    fdr    = {"SITE1;SITE2;": 0.001}

    result = _run(empty_wb, probs=probs, ratios=ratios, fdr=fdr)
    assert "SITE1" in result["KinaseX"]
    assert "SITE2" in result["KinaseX"]


def test_split_no_incluye_strings_vacios(empty_wb):
    """
    the split by ';' of 'SITE1;' produces ['SITE1', ''].
    the empty string cannot be included in the susbtrate list
    """
    probs  = {"KinaseX": {"SITE1;": 1.0}}
    ratios = {"KinaseX": {"SITE1;": 1.0}}
    fdr    = {"SITE1;": 0.001}

    result = _run(empty_wb, probs=probs, ratios=ratios, fdr=fdr)
    assert "" not in result["KinaseX"]
    assert "SITE1" in result["KinaseX"]


# ==== TESTS: KINASE WITHOUT SUBSTRATES ====

def test_kinasa_sin_sitios_validos_tiene_lista_vacia(empty_wb):
    #KinaseY has an empty dict of probabilities : empty list of substrates
    result = _run(empty_wb)
    assert result["KinaseY"] == []


# ==== TESTS: CREATED SHEET ====

def test_crea_hoja_PutativeKinaseSubstrates(empty_wb):
    #Function must create the sheet 'PutativeKinaseSubstrates'
    _run(empty_wb)
    assert "PutativeKinaseSubstrates" in empty_wb.sheetnames


def test_hoja_PutativeKinaseSubstrates_tiene_header_correcto(empty_wb):
    #Sheet must have headers: 'kinase', 'n', 'substrates'
    _run(empty_wb)
    ws = empty_wb["PutativeKinaseSubstrates"]
    header = [cell.value for cell in list(ws.iter_rows(values_only=False))[0]]
    assert header[0] == "kinase"
    assert header[1] == "n"
    assert header[2] == "substrates"

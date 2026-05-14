"""
test_getDataForKusterCompounds.py

unit test for the function getDataForKusterCompounds().

This function reads an excel workbook and extract:
  - fValues: fold changes : {fosfositio: {compuesto: valor}}
  - pValues: p-values : {fosfositio: {compuesto: valor}}
  - fdrValues: FDR : {fosfositio: valor}
  - compoundsCellLine: maps short compound : full columns name
  - sitesCellLine: organizer list of phosphosites

Run:
    pytest tests/python/test_getDataForKusterCompounds.py -v
"""

import pytest
import openpyxl
from conftest import make_workbook
from ebdtFunctions import getDataForKusterCompounds


# ==== WORKBOOK DE PRUEBA ====

@pytest.fixture
def wb_simple():
    """
    Workbook min with 2 compounds and 2 psites.

    Sheet 'fold':
        sh.index.sites | FDR  | MCF7.CompA.fold | MCF7.CompB.fold
        SITE1;         | 0.01 | -2.5            | 1.3
        SITE2;         | 0.50 |  0.8            | -0.3

    Sheet 'pvalue':
        sh.index.sites | None | MCF7.CompA.p.value | MCF7.CompB.p.value
        SITE1;         | 0.01 | 0.010              | 0.300
        SITE2;         | 0.50 | 0.800              | 0.900
    """
    return make_workbook(
        fold_rows=[
            ["sh.index.sites", "FDR",  "MCF7.CompA.fold",    "MCF7.CompB.fold"],
            ["SITE1;",          0.01,  -2.5,                  1.3],
            ["SITE2;",          0.50,   0.8,                 -0.3],
        ],
        pvalue_rows=[
            ["sh.index.sites", None,   "MCF7.CompA.p.value", "MCF7.CompB.p.value"],
            ["SITE1;",          0.01,   0.010,                0.300],
            ["SITE2;",          0.50,   0.800,                0.900],
        ],
    )


# ==== HELPER ====

def _run(wb):
    #runs getDataForKusterCompounds with the parametres of the simple workbook
    return getDataForKusterCompounds(wb, ["CompA", "CompB"], numCompounds=2)


# ==== TESTS: PHOSPHOSITES ====

def test_extrae_lista_de_sitios(wb_simple):
    #sitesCellLine must contain the names of all the organized psites
    _, _, _, _, sites = _run(wb_simple)
    assert sites == ["SITE1;", "SITE2;"]


def test_numero_de_sitios_correcto(wb_simple):
    #number of psites must agree with the data rows
    _, _, _, _, sites = _run(wb_simple)
    assert len(sites) == 2


#==== TESTS: FOLD VALUES ====

def test_extrae_fold_values_correctamente(wb_simple):
    #fValues[site][compound] must contains the numerical value of the fold change
    fValues, _, _, _, _ = _run(wb_simple)

    assert fValues["SITE1;"]["CompA"] == pytest.approx(-2.5)
    assert fValues["SITE1;"]["CompB"] == pytest.approx(1.3)
    assert fValues["SITE2;"]["CompA"] == pytest.approx(0.8)
    assert fValues["SITE2;"]["CompB"] == pytest.approx(-0.3)


def test_claves_fvalues_coinciden_con_sites(wb_simple):
    #The keys in fValues must be exactly the same as the sitesCellLine keys
    fValues, _, _, _, sites = _run(wb_simple)
    assert set(fValues.keys()) == set(sites)


def test_fvalues_son_numericos(wb_simple):
    #values of fValues must be float, not strings
    fValues, _, _, _, _ = _run(wb_simple)
    for site, compounds in fValues.items():
        for compound, val in compounds.items():
            assert isinstance(val, float), (
                f"fValues[{site}][{compound}] = {val!r} no es float"
            )


# ==== TESTS: P-VALUES ====

def test_extrae_pvalues_correctamente(wb_simple):
    #pValues[site][compound] must contain the right p-value
    _, pValues, _, _, _ = _run(wb_simple)

    assert pValues["SITE1;"]["CompA"] == pytest.approx(0.010)
    assert pValues["SITE1;"]["CompB"] == pytest.approx(0.300)
    assert pValues["SITE2;"]["CompA"] == pytest.approx(0.800)


def test_claves_pvalues_coinciden_con_fvalues(wb_simple):
    #fValues and pValues must have exactly the same psites and compounds
    fValues, pValues, _, _, _ = _run(wb_simple)
    assert set(fValues.keys()) == set(pValues.keys())
    for site in fValues:
        assert set(fValues[site].keys()) == set(pValues[site].keys())


# ==== TESTS: FDR ====

def test_extrae_fdr_correctamente(wb_simple):
    #fdrValues[site] must contain the correct FDR value
    _, _, fdrValues, _, _ = _run(wb_simple)

    assert fdrValues["SITE1;"] == pytest.approx(0.01)
    assert fdrValues["SITE2;"] == pytest.approx(0.50)


def test_fdr_none_se_convierte_a_cero():
    #if FDR cell is None, it must be treated as 0 (do not throw error)
    wb = make_workbook(
        fold_rows=[
            ["sh.index.sites", "FDR", "MCF7.CompA.fold"],
            ["SITE1;",          None,  -2.5],
        ],
        pvalue_rows=[
            ["sh.index.sites", None, "MCF7.CompA.p.value"],
            ["SITE1;",          None,  0.01],
        ],
    )
    _, _, fdrValues, _, _ = getDataForKusterCompounds(wb, ["CompA"], 1)

    assert fdrValues["SITE1;"] == pytest.approx(0.0)


def test_fdr_string_vacio_se_convierte_a_cero():
    #if FDR cell is an empty string, it must be trated as 0
    wb = make_workbook(
        fold_rows=[
            ["sh.index.sites", "FDR", "MCF7.CompA.fold"],
            ["SITE1;",          "",    -2.5],
        ],
        pvalue_rows=[
            ["sh.index.sites", None, "MCF7.CompA.p.value"],
            ["SITE1;",          "",    0.01],
        ],
    )
    _, _, fdrValues, _, _ = getDataForKusterCompounds(wb, ["CompA"], 1)

    assert fdrValues["SITE1;"] == pytest.approx(0.0)


# ==== TESTS: COMPOUNDSCELLLINE ====

def test_compoundsCellLine_keys_son_nombres_cortos(wb_simple):
    #Keys of compoundsCellLine must be short names (without prefix or suffix)
    _, _, _, compoundsCellLine, _ = _run(wb_simple)

    assert "CompA" in compoundsCellLine
    assert "CompB" in compoundsCellLine


def test_compoundsCellLine_valores_son_nombres_completos(wb_simple):
    #values of compoundsCellLine must be the complete column names
    _, _, _, compoundsCellLine, _ = _run(wb_simple)

    assert compoundsCellLine["CompA"] == "MCF7.CompA.fold"
    assert compoundsCellLine["CompB"] == "MCF7.CompB.fold"


# ==== TESTS: CREATED SHEETS ====

def test_crea_hoja_pvalue_select(wb_simple):
    #Function must create the sheet 'pvalue.select' if it does not exist
    _run(wb_simple)
    assert "pvalue.select" in wb_simple.sheetnames


def test_crea_hoja_fold_select(wb_simple):
    #function must creathe the sheet 'fold.select' if it does not exist
    _run(wb_simple)
    assert "fold.select" in wb_simple.sheetnames


def test_no_duplica_hojas_si_ya_existen(wb_simple):
    #running the fuction 2 times should not duplicate the sheet
    _run(wb_simple)
    _run(wb_simple)  #second run
    assert wb_simple.sheetnames.count("pvalue.select") == 1
    assert wb_simple.sheetnames.count("fold.select") == 1

"""
export_PDTs.py

Reads the PutativeKinaseSubstrates golden CSVs for each cell line,
filters out kinases with no substrates (n=0), and writes one Excel
workbook with one sheet per cell line plus a "Combined" sheet:
    A = kinase
    B = n (number of PDTs)
    C = substrates (semicolon-separated)

The Combined sheet contains the union of substrates for each kinase
across all three cell lines (unique PDTs only, sorted).

Run from the project root:
    python export_PDTs.py
"""

import csv
import os
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font

GOLDEN_DIR = os.path.join("tests", "golden")
CELL_LINES = ["MCF7", "HL60", "NTERA2"]

output_path = "PDTs.xlsx"
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default empty sheet

# dict to accumulate substrates across cell lines: {kinase: set of substrates}
combined: dict[str, set] = defaultdict(set)


def _write_sheet(ws, data_rows):
    """Write header + data rows to a worksheet. Returns number of data rows written."""
    ws["A1"] = "kinase"
    ws["B1"] = "n"
    ws["C1"] = "substrates"
    for cell in [ws["A1"], ws["B1"], ws["C1"]]:
        cell.font = Font(bold=True)

    for excel_row, (kinase, n, substrates) in enumerate(data_rows, start=2):
        ws.cell(row=excel_row, column=1).value = kinase
        ws.cell(row=excel_row, column=2).value = n
        ws.cell(row=excel_row, column=3).value = substrates
    return len(data_rows)


for cell_line in CELL_LINES:
    input_path = os.path.join(GOLDEN_DIR, f"{cell_line}_PutativeKinaseSubstrates.csv")

    with open(input_path, "rt", encoding="utf-8") as f:
        rows = list(csv.reader(f))

    ws = wb.create_sheet(title=cell_line)
    data_rows = []

    for row in rows[1:]:
        if not row:
            continue
        kinase = row[0]
        n      = int(float(row[1]))
        subs   = row[2] if len(row) > 2 else ""
        if n == 0:
            continue
        data_rows.append((kinase, n, subs))

        # accumulate into combined dict
        for site in subs.split(";"):
            site = site.strip()
            if site:
                combined[kinase].add(site)

    count = _write_sheet(ws, data_rows)
    print(f"{cell_line}: {count} kinases with PDTs")

# ── Combined sheet ─────────────────────────────────────────────────────────────
ws_comb = wb.create_sheet(title="Combined")
combined_rows = []
for kinase in sorted(combined):
    sites = sorted(combined[kinase])
    subs_str = ";".join(sites)
    combined_rows.append((kinase, len(sites), subs_str))

count_comb = _write_sheet(ws_comb, combined_rows)
print(f"Combined: {count_comb} kinases with PDTs "
      f"({sum(len(v) for v in combined.values())} total unique PDTs)")

wb.save(output_path)
print(f"Saved -> {output_path}")
